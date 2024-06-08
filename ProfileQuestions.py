import openpyxl
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox
from rake_nltk import Rake
import nltk
import os

# This thing is a need
nltk.download('stopwords')

questions = [
    "What is your name?",
    "How old are you?",
    "What is your preferred field?",
    "What is your highest qualification/education?",
    "How many years of experience do you have in this field? (if none, write 'none')",
    "Write down your skills."
]


def detect_key_phrases(text):
    rake = Rake()
    rake.extract_keywords_from_text(text)
    key_phrases = rake.get_ranked_phrases()
    return key_phrases

def write_to_excel(all_responses):
    file_path = 'chatbot_responses.xlsx'

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Field", "Highest qualification", "Years of experience", "Skills"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active
    if len(all_responses) > 5:
        key_phrases = detect_key_phrases(all_responses[5])
        unique_key_phrases = ', '.join(set(key_phrases))
        all_responses[5] = unique_key_phrases

    ws.append(all_responses)
    wb.save(file_path)
    print(f"Data written to {file_path}")

def submit_form():
    #This combines the GUI, gets the data from the field, aka (xxx)_entry
    name = name_entry.get()
    age = age_entry.get()
    field = field_entry.get()
    qualification = qualification_entry.get()
    experience = experience_entry.get()
    skills = skills_entry.get()

    # GUI responses to data in here
    user_responses = [name, age, field, qualification, experience, skills]
    write_to_excel(user_responses)
    messagebox.showinfo("Success", "Profile Made!")


#TKinter sectio

questionnaire = Tk()
questionnaire.title("Questionnaire")
questionnaire.geometry('400x400')

#name
name_label = Label(questionnaire, text=questions[0])
name_label.pack()
name_entry = Entry(questionnaire)
name_entry.pack()
#age
age_label = Label(questionnaire, text=questions[1])
age_label.pack()
age_entry = Entry(questionnaire)
age_entry.pack()
#field
field_label = Label(questionnaire, text=questions[2])
field_label.pack()
field_entry = Entry(questionnaire)
field_entry.pack()
#qualification
qualification_label = Label(questionnaire, text=questions[3])
qualification_label.pack()
qualification_entry = Entry(questionnaire)
qualification_entry.pack()
#experience
experience_label = Label(questionnaire, text=questions[4])
experience_label.pack()
experience_entry = Entry(questionnaire)
experience_entry.pack()
#skills
skills_label = Label(questionnaire, text=questions[5])
skills_label.pack()
skills_entry = Entry(questionnaire)
skills_entry.pack()
#will add and modify later to reflect actualy questions to ask
submit_button = Button(questionnaire, text="Submit", command=submit_form)
submit_button.pack()

questionnaire.mainloop()
