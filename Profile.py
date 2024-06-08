import os
import re
import tkinter as tk
from tkinter import font as tkFont
from openpyxl import load_workbook


def read_excel_file():
    file_path = "chatbot_responses.xlsx"
    if not os.path.exists(file_path):
        print("Excel file not found!")
        return None

    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    last_row = [cell.value for cell in ws[ws.max_row]]

    # Print headers and last_row for debugging
    print("Headers:", headers)
    print("Last Row:", last_row)

    return dict(zip(headers, last_row))


def read_latest_resume():
    resume_files = [f for f in os.listdir() if f.startswith("resume_") and f.endswith(".txt")]
    if not resume_files:
        print("No resume files found!")
        return None

    # Extract the number from each file and find the highest number
    latest_resume = max(resume_files, key=lambda x: int(x.split('_')[1].split('.')[0]))
    with open(latest_resume, "r") as file:
        resume_content = file.read()

    email = extract_detail(resume_content, r'Email: (.+)')
    phone = extract_detail(resume_content, r'Phone: (.+)')
    address = extract_detail(resume_content, r'Address: (.+)')
    degrees = extract_detail(resume_content, r'Degrees: (.+)')

    return resume_content, email, phone, address, degrees


def extract_detail(text, pattern):
    match = re.search(pattern, text)
    return match.group(1) if match else "Not Available"


def show_profile(profile_data, resume_content, email, phone, address, degrees):
    root = tk.Tk()
    root.title("Profile")

    # Set the popup window to appear at the center of the root window
    root.geometry("400x400")
    root.resizable(False, False)

    # Define font style and size
    font_style = tkFont.Font(family="Helvetica", size=10)

    profile_text = (f"\\\\\\\\\n"
                    f"\\\\\\\\        {profile_data['Name']}\n"
                    f"\\\\\\\\\n"
                    f"\\\\\\\\        {email}\n"
                    f"\\\\\\\\        {phone}\n"
                    f"\\\\\\\\        {address}\n"
                    f"\\\\\\\\\n"
                    f"SKILLS: {profile_data['Skills']}\n"
                    f"DEGREES: {degrees}\n"
                    f"HIGHEST QUALIFICATION: {profile_data['Highest qualification']}\n"
                    f"CURRENLTY LOOKING TO WORK IN FIELD: {profile_data['Field']}\n"
                    f"YEARS OF EXPERIENCE: {profile_data['Years of experience']}\n")

    profile_label = tk.Label(root, text=profile_text, justify="left", wraplength=380, font=font_style)
    profile_label.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    resume_button = tk.Button(root, text="View Resume", command=lambda: show_resume_popup(resume_content))
    resume_button.pack(pady=10)

    root.mainloop()


def show_resume_popup(resume_content):
    popup = tk.Toplevel()
    popup.title("Generated Resume")

    # Set the popup window to appear at the center of the root window
    popup.geometry("400x400")
    popup.resizable(False, False)

    # Define font style and size
    font_style = tkFont.Font(family="Helvetica", size=10)

    resume_label = tk.Label(popup, text=resume_content, justify="left", wraplength=380, font=font_style)
    resume_label.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    ok_button = tk.Button(popup, text="OK", command=popup.destroy)
    ok_button.pack(pady=10)

    # Add Confirm button to save the resume
    confirm_button = tk.Button(popup, text="Confirm", command=lambda: save_confirmed_resume(resume_content))
    confirm_button.pack(pady=10)


def save_confirmed_resume(resume_content):
    # Create a new file to save the confirmed resume
    with open("confirmed_resume.txt", "w") as file:
        file.write(resume_content)
    print("Resume saved as confirmed_resume.txt")


def main():
    profile_data = read_excel_file()
    resume_content, email, phone, address, degrees = read_latest_resume()

    # Print profile data for debugging
    print("Profile Data:", profile_data)

    if profile_data and resume_content:
        show_profile(profile_data, resume_content, email, phone, address, degrees)
    else:
        print("Unable to display profile.")


if __name__ == "__main__":
    main()
